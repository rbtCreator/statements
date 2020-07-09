# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'C:\\Users\\gulya\\Desktop\\sessions\\designs\\loading.ui'
#
# Created by: PyQt5 UI code generator 5.12.2
#
# WARNING! All changes made in this file will be lost!

#import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook, load_workbook


class Ui_MainWindow(object):
    def __init__(self, SV_FILENAME, COURSE, SESSIONS_FILENAME, SEMESTER):
        super().__init__()

        self.mainwindow = QtWidgets.QMainWindow()
        self.setupUi(self.mainwindow)
        self.mainwindow.show()
        self.SV_FILENAME = SV_FILENAME
        self.COURSE = COURSE
        self.SESSIONS_FILENAME = SESSIONS_FILENAME
        self.SEMESTER = SEMESTER
        self.parse()

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(496, 218)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.l_choose_SV = QtWidgets.QLabel(self.centralwidget)
        self.l_choose_SV.setGeometry(QtCore.QRect(150, 60, 271, 71))
        self.l_choose_SV.setObjectName("l_choose_SV")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.l_choose_SV.setText(_translate("MainWindow", "Пожалуйста, подождите..."))

    def parse(self):
        wb_SV = load_workbook(self.SV_FILENAME)
        SV_sheetnames = wb_SV.sheetnames
        wb_SESSIONS = load_workbook(self.SESSIONS_FILENAME)

        # for study groups...
        for sheet in wb_SESSIONS:
            #if there's a conformity among SV and SESSIONS sheets...
            if self.COURSE in sheet.title and sheet.title in SV_sheetnames:
                
                # FILLING ITSELF
                ws_group_session = sheet
                ws_group_sv = wb_SV[sheet.title]

                # getting list of disciplines from SESSIONS
                disciplines = []
                r = 5
                c = 6
                breakpoint = 0
                while True:
                    val = ws_group_session.cell(row=r, column=c).value
                    if type(val) != str:
                        breakpoint += 1
                        if breakpoint > 20:
                            break
                    disciplines.append(str(val))
                    c += 1
                    breakpoint = 0


                # SDM is 'surname - discipline - mark'
                sdm_dict = {}
                # getting list of surnames
                c = 4
                for r in range(6, 40):
                    val = ws_group_session.cell(row=r, column=c).value
                    if type(val) == str:
                        if len(val) > 3:
                            val = val.split()[0]
                            sdm_dict[val] = dict()

                            c_m = 6
                            for discipline in disciplines:
                                mark = ws_group_session.cell(row=r, column=c_m).value

                                sdm_dict[val][discipline] = mark
                                c_m += 1
                            
                            print(sdm_dict)

                # filling the SV
                r = 4
                c = 3
                val = int(ws_group_sv.cell(row=r, column=c).value)
                '''
                while True:
                    if val == self.COURSE:
                        # get the discipline name
                        discipline = str(ws_group_sv.cell(row=r+1, column=c).value)
'''
                    
                

                

        


'''
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
'''